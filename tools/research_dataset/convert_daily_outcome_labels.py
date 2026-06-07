#!/usr/bin/env python3
"""Convert Sookta daily outcome-label workbooks into training/test fixtures.

The workbook is expected to contain a `Training_Data` sheet where each row is
one farmer-level 7-transaction window. Features are the normalized values used
by `DailyInjuryPredictionService`; the target is
`requires_medical_treatment_within_7_days` with values 0 or 1.
"""

from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FEATURE_COLUMNS = [
    "avg_score_before_norm",
    "max_score_before_norm",
    "avg_score_after_norm",
    "high_or_above_days_norm",
    "very_high_days_norm",
    "no_improvement_days_norm",
    "trunk_high_days_norm",
    "neck_or_upper_limb_high_days_norm",
    "iso_days_norm",
    "avg_economic_loss_norm",
    "repeated_same_activity_norm",
    "recent_score_slope_norm",
]

TARGET_COLUMN = "requires_medical_treatment_within_7_days"

METADATA_COLUMNS = [
    "row_type",
    "window_id",
    "farmer_id",
    "participant_code",
    "window_start_date",
    "window_end_date",
    "transaction_count",
    "transaction_ids",
    "activity_summary",
    "assessment_methods",
    "medical_visit_within_7_days",
    "treatment_required_within_7_days",
    "msd_symptom_present",
    "msd_symptom_location",
    "msd_symptom_severity",
    "lost_workdays_7d",
    "direct_medical_cost_thb",
    "productivity_loss_thb",
    "label_confidence",
    "outcome_source",
    "reviewer_id",
    "reviewed_at",
    "notes",
]

REQUIRED_COLUMNS = [
    "window_id",
    "farmer_id",
    "transaction_count",
    *FEATURE_COLUMNS,
    TARGET_COLUMN,
]


class ConversionError(ValueError):
    """Raised when workbook data cannot be converted safely."""


@dataclass(frozen=True)
class ConvertedOutcomeDataset:
    rows: list[dict[str, Any]]
    errors: list[str]


def convert_workbook(
    workbook_path: Path,
    *,
    sheet_name: str = "Training_Data",
    include_examples: bool = False,
) -> ConvertedOutcomeDataset:
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ConversionError(
            f"Workbook must contain a '{sheet_name}' sheet. "
            f"Found: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    headers = [
        _normalize_header(value)
        for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    _validate_headers(headers)

    rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for excel_row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        raw = {
            header: _clean_cell(value)
            for header, value in zip(headers, values)
            if header
        }
        if _is_blank_row(raw):
            continue
        if str(raw.get("row_type", "")).strip().lower() == "example" and not include_examples:
            continue

        try:
            rows.append(_convert_row(raw, excel_row_number))
        except ConversionError as error:
            errors.append(str(error))

    return ConvertedOutcomeDataset(rows=rows, errors=errors)


def write_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = [*METADATA_COLUMNS, *FEATURE_COLUMNS, TARGET_COLUMN]
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema": "sookta_daily_injury_outcome_labels_v1",
        "target": TARGET_COLUMN,
        "featureColumns": FEATURE_COLUMNS,
        "rowCount": len(rows),
        "rows": rows,
    }
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _validate_headers(headers: list[str]) -> None:
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ConversionError(f"Training_Data sheet is missing columns: {', '.join(missing)}")


def _convert_row(raw: dict[str, Any], excel_row_number: int) -> dict[str, Any]:
    missing_required = [
        column
        for column in REQUIRED_COLUMNS
        if raw.get(column) in (None, "")
    ]
    if missing_required:
        raise ConversionError(
            f"Row {excel_row_number}: missing required values: "
            f"{', '.join(missing_required)}"
        )

    transaction_count = _as_int(raw["transaction_count"], excel_row_number, "transaction_count")
    if transaction_count != 7:
        raise ConversionError(
            f"Row {excel_row_number}: transaction_count must be 7; got {transaction_count}"
        )

    target = _as_int(raw[TARGET_COLUMN], excel_row_number, TARGET_COLUMN)
    if target not in (0, 1):
        raise ConversionError(
            f"Row {excel_row_number}: {TARGET_COLUMN} must be 0 or 1; got {target}"
        )

    row: dict[str, Any] = {}
    for column in METADATA_COLUMNS:
        value = raw.get(column, "")
        if column == "transaction_count" and value != "":
            row[column] = transaction_count
        elif column in {
            "medical_visit_within_7_days",
            "treatment_required_within_7_days",
            "msd_symptom_present",
        } and value != "":
            row[column] = _as_int(value, excel_row_number, column)
        elif column in {
            "lost_workdays_7d",
            "direct_medical_cost_thb",
            "productivity_loss_thb",
        } and value != "":
            row[column] = _as_float(value, excel_row_number, column)
        else:
            row[column] = value

    for column in FEATURE_COLUMNS:
        value = _as_float(raw[column], excel_row_number, column)
        if value < 0 or value > 1:
            raise ConversionError(
                f"Row {excel_row_number}: {column} must be within 0..1; got {value}"
            )
        row[column] = value

    row[TARGET_COLUMN] = target
    return row


def _as_int(value: Any, excel_row_number: int, column: str) -> int:
    try:
        if isinstance(value, str):
            value = value.strip()
        return int(float(value))
    except (TypeError, ValueError):
        raise ConversionError(f"Row {excel_row_number}: {column} must be an integer")


def _as_float(value: Any, excel_row_number: int, column: str) -> float:
    try:
        if isinstance(value, str):
            value = value.strip()
        return float(value)
    except (TypeError, ValueError):
        raise ConversionError(f"Row {excel_row_number}: {column} must be numeric")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def _is_blank_row(row: dict[str, Any]) -> bool:
    return all(value in ("", None) for value in row.values())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Outcome label .xlsx workbook")
    parser.add_argument(
        "--sheet",
        default="Training_Data",
        help="Worksheet name to read",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=Path("data/research/extracted/daily_injury_outcome_training_ready.csv"),
        help="CSV output path",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path("data/research/extracted/daily_injury_outcome_training_ready.json"),
        help="JSON fixture output path",
    )
    parser.add_argument(
        "--include-examples",
        action="store_true",
        help="Include rows marked row_type=example",
    )
    args = parser.parse_args()

    dataset = convert_workbook(
        args.workbook,
        sheet_name=args.sheet,
        include_examples=args.include_examples,
    )
    if dataset.errors:
        for error in dataset.errors:
            print(error)
        return 2

    write_csv(dataset.rows, args.output_csv)
    write_json(dataset.rows, args.output_json)
    print(f"Converted rows: {len(dataset.rows)}")
    print(f"CSV: {args.output_csv}")
    print(f"JSON: {args.output_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
