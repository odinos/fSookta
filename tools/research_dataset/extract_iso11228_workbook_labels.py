#!/usr/bin/env python3
"""Extract ISO11228 session labels from the research workbook."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ACTIVITY_MAP = {
    "transplanting": "transplanting",
    "fertilizing": "fertilizing",
    "pesticide spraying": "pesticide_spraying",
    "pruning": "pruning",
    "harvesting": "harvesting",
    "on-farm transport": "on_farm_transport",
}

SUMMARY_KEYS = {
    "A. ท่าทางลำตัวและคอ": "A_trunk_neck",
    "B. แขน ไหล่ มือ": "B_arm_shoulder_hand",
    "C. แรงที่ใช้": "C_force",
    "D. ความถี่และระยะเวลา": "D_frequency_duration",
    "E. ระยะเอื้อม/ตำแหน่งเครื่องมือ": "E_reach_tool_position",
    "F. สภาพแวดล้อม": "F_environment",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract ISO11228 labels from ISO11228 Pirawan.xlsx.",
    )
    parser.add_argument(
        "--input",
        default="/Users/kpc/Documents/Doc/ISO11228 Pirawan.xlsx",
        help="Research ISO11228 workbook path.",
    )
    parser.add_argument(
        "--output",
        default="data/research/expert_labels/iso11228_expert_labels.csv",
        help="Canonical CSV consumed by the training pipeline.",
    )
    parser.add_argument(
        "--training-ready-output",
        default="data/research/expert_labels/iso11228_training_ready_20260603.csv",
        help="Compatibility copy used by older docs and audit bundles.",
    )
    return parser.parse_args()


def clean(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def canonical_session(raw: str) -> str:
    return re.sub(r"^A", "", raw.strip(), flags=re.IGNORECASE)


def canonical_activity(values: list[str]) -> str:
    for value in values:
        normalized = value.strip().lower()
        if normalized in ACTIVITY_MAP:
            return ACTIVITY_MAP[normalized]
    return ""


def numeric(value: object | None) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def risk_from_total(total: float, worksheet_risk: str) -> str:
    if worksheet_risk:
        return worksheet_risk
    if total <= 9:
        return "ต่ำ"
    if total <= 13:
        return "ปานกลาง"
    return "สูง"


def extract_rows(input_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    output: list[dict[str, str]] = []
    index = 0
    while index < len(rows):
        values = [clean(value) for value in rows[index]]
        first = values[0] if values else ""
        if first and "ขั้นตอนการทำงาน:" in values:
            session_id = canonical_session(first)
            activity = canonical_activity(values[2:5])
            has_media_note = any("ไม่มีวิดีโอ" in value for value in values)
            scores = {key: 0 for key in SUMMARY_KEYS.values()}
            total = 0.0
            risk = ""
            cursor = index + 1
            while cursor < len(rows):
                line = [clean(value) for value in rows[cursor]]
                next_first = line[0] if line else ""
                if next_first and "ขั้นตอนการทำงาน:" in line:
                    break
                label = line[1] if len(line) > 1 else ""
                if label in SUMMARY_KEYS:
                    scores[SUMMARY_KEYS[label]] = int(numeric(line[2]))
                if label == "สรุป":
                    total = numeric(line[2])
                    risk = line[5] if len(line) > 5 else ""
                cursor += 1

            if activity and session_id and total > 0 and not has_media_note:
                output.append(
                    {
                        "source_file": input_path.name,
                        "activity": activity,
                        "session_id": session_id,
                        "iso11228_total_score": f"{total:.0f}",
                        "risk_level_th": risk_from_total(total, risk),
                        **{key: str(value) for key, value in scores.items()},
                        "label_source": "research_team_iso11228_workbook_full_20260607",
                    }
                )
            index = cursor
        else:
            index += 1
    return output


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "source_file",
        "activity",
        "session_id",
        "iso11228_total_score",
        "risk_level_th",
        "A_trunk_neck",
        "B_arm_shoulder_hand",
        "C_force",
        "D_frequency_duration",
        "E_reach_tool_position",
        "F_environment",
        "label_source",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    rows = extract_rows(Path(args.input))
    write_csv(Path(args.output), rows)
    write_csv(Path(args.training_ready_output), rows)
    print(f"extracted={len(rows)} output={args.output}")


if __name__ == "__main__":
    main()
